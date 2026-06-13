#!/usr/bin/env python3
"""Create the two-sheet autocite audit workbook."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SHEETS = ("Excluded Comments", "Resolved Comments")


def atomic_save_workbook(wb: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp.xlsx")
    wb.save(tmp)
    tmp.replace(path)


def load_rows(path: Path | None) -> list[dict[str, Any]]:
    if path is None or not path.exists():
        return []
    if path.suffix.lower() == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, list):
            return [dict(row) for row in data]
        for key in ("excluded", "resolved", "rows", "comments", "mappings"):
            if isinstance(data.get(key), list):
                return [dict(row) for row in data[key]]
    with path.open(encoding="utf-8", newline="") as fh:
        return list(csv.DictReader(fh))
    raise ValueError(f"Unsupported input: {path}")


def columns(rows: list[dict[str, Any]], preferred: list[str]) -> list[str]:
    rest = sorted({key for row in rows for key in row} - set(preferred))
    return preferred + rest


def write_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]], preferred: list[str]) -> None:
    ws = wb.create_sheet(title)
    cols = columns(rows, preferred) or preferred
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([row.get(col, "") for col in cols])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for index, col in enumerate(cols, start=1):
        width = min(80, max(12, max(len(str(ws.cell(row=r, column=index).value or "")) for r in range(1, ws.max_row + 1)) + 2))
        ws.column_dimensions[get_column_letter(index)].width = width
        if col.lower() == "confidence":
            letter = get_column_letter(index)
            rng = f"{letter}2:{letter}{max(ws.max_row, 2)}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"high"'], fill=PatternFill("solid", fgColor="9DC3E6")))
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"moderate"'], fill=PatternFill("solid", fgColor="FFD966")))
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"low"'], fill=PatternFill("solid", fgColor="F4B183")))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excluded", type=Path)
    parser.add_argument("--resolved", type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    excluded = load_rows(args.excluded)
    resolved = load_rows(args.resolved)
    for row in excluded:
        row.setdefault("confidence", "low")
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(
        wb,
        "Excluded Comments",
        excluded,
        ["comment_id", "author", "date", "anchor_text", "paragraph_text", "comment_text", "confidence", "exclusion_reason"],
    )
    write_sheet(
        wb,
        "Resolved Comments",
        resolved,
        [
            "comment_id",
            "author",
            "date",
            "anchor_text",
            "paragraph_text",
            "comment_text",
            "resolved_references",
            "confidence",
            "insertion_text",
            "record_numbers",
            "source_method",
            "agent_agreement_status",
        ],
    )
    atomic_save_workbook(wb, args.output)
    print(json.dumps({"output": str(args.output), "sheets": SHEETS, "excluded_rows": len(excluded), "resolved_rows": len(resolved)}, indent=2))


if __name__ == "__main__":
    main()
